# Baseball neural net data collection
#
# ### All data is coming from Baseball-reference.com
#
# This script has two modes, -players and -teams
# -players collects game info as a list of players
# -teams collects game info as stats about the teams
# 
# ### -teams ###
# First we just need to get the number of the team in alphabetical order
# (every team will be represented by a number in the range [0, 29])
# This will be used to fetch the input for the neural network
# away_id = number of away team in alphabetical ordering of MLB teams
# home_id = number of home team in alphabetical ordering of MLB teams
#
# Later, we can add some of this other data:
# Data for each game (x2: home/away)
# (We can go for more later, but these are 5 good ones to start with)
# --------------
# 1. Starting pitcher ERA
# 2. Starting pitcher WHIP
# 3. Team OBP
# 4. Average runs scored
# 5. Team RBIs
# --------------
# Input data will look like this:
# [away_era, away_whip, away_obp, away_ars, away_rbi,
#  home_era, home_whip, home_obp, home_ars, home_rbi,
#  away_id, home_id]
#
# Output (solution) data will look like this:
# [away_team_win, home_team_win]
# (i.e., [1, 0] or [0, 1])
# --------------
# Possible additional data points:
# 1. Team overall record
# 2. Starting pitcher's record against opposing team
# 3. Team's streak (e.g., W3, L1, etc.)
#
# ### -players ###
# Every game is encoded as a list of players
# Every player is assigned a numerical value
# The game is written to a csv file as a list of numbers
# which correspond to the players that started for each team in that game
# This mode is simpler but the collection takes a little longer (?)
##################################################

from requests import get
from bs4 import BeautifulSoup
from collections import defaultdict
import re
from openpyxl import Workbook
import html5lib
import pickle
from sklearn import preprocessing
import numpy as np
import csv

base_url = "https://baseball-reference.com"

def get_id(team):
	# to help with one-hot embeddings
	teams_alpha = {"Arizona Diamondbacks": 0, "Atlanta Braves": 1, "Baltimore Orioles": 2,
				   "Boston Red Sox": 3, "Chicago Cubs" : 4, "Chicago White Sox": 5,
				   "Cincinnati Reds": 6, "Cleveland Indians": 7, "Colorado Rockies": 8,
				   "Detroit Tigers": 9, "Houston Astros": 10, "Kansas City Royals": 11,
				   "Los Angeles Angels": 12, "Los Angeles Angels of Anaheim": 12,
				   # Angels name change on BR.com between 2016 and 2017
				   "Los Angeles Dodgers": 13, "Miami Marlins": 14,
				   "Milwaukee Brewers": 15, "Minnesota Twins": 16, "New York Mets": 17,
				   "New York Yankees": 18, "Oakland Athletics": 19, "Philadelphia Phillies": 20,
				   "Pittsburgh Pirates": 21, "San Diego Padres": 22, "San Francisco Giants": 23,
				   "Seattle Mariners": 24, "St. Louis Cardinals": 25, "Tampa Bay Rays": 26,
				   "Texas Rangers": 27, "Toronto Blue Jays": 28, "Washington Nationals": 29}
	return teams_alpha[team]

class Team():
	def __init__(self, name):
		self.name = name
		self.games = 0 # games played
		self.runs = 0 # runs scored
		self.rbis = 0

	def avg_runs(self):
		if self.games == 0:
			return 0 # what should this (default) value be?
		return (self.runs * 1.0) / self.games

	def __str__(self): # overload print()
		result = "### Team: %s ###\n" % self.name +\
				 "# Games played: %d\n" % self.games +\
				 "# Runs scored: %d\n" % self.runs +\
				 "# RBIs: %d" % self.rbis
		return result

class Pitcher():
	def __init__(self, name):
		self.name = name
		self.bb = 0 # bases on balls
		self.hits = 0 # hits allowed
		self.er = 0 # earned runs
		self.ip = 0 # innings pitched

	def whip(self):
		if self.ip == 0:
			return 0 # what should this (default) value be?
		return (self.bb + self.hits + 0.0)/ self.ip

	def era(self):
		if self.ip == 0:
			return 0 # what should this (default) value be?
		return ((self.er + 0.0)/self.ip) * 9

	def __str__(self): # overload print()
		result = "### Pitcher: %s ###\n" % self.name +\
				 "# Bases on balls: %d\n" % self.bb +\
				 "# Hits: %d\n" % self.hits +\
				 "# Earned runs: %d\n" % self.er +\
				 "# Innings pitched: %.1f\n" % self.ip +\
				 "# WHIP: %.4f\n" % self.whip() +\
				 "# ERA: %.2f" % self.era()
		return result

class Batter():
	def __init__(self, name):
		self.name = name
		self.obp = 0.0 # what should this (default) value be?
		# self.team = None

	def __str__(self): # overload print
		result = "### Batter: %s ###\n" % self.name +\
				 "# OBP: %.3f" % self.obp
		return result

class keydefaultdict(defaultdict):
	# to pass key to Team class constructor if team not in dict
	# thanks to Jochen Ritzel on Stack Overflow
    def __missing__(self, key):
        if self.default_factory is None:
            raise KeyError( key )
        else:
            ret = self[key] = self.default_factory(key)
            return ret

def save_obj(obj, name):
	# to save dictionaries in case of program stopping mid-way through 
    with open('obj/'+ name + '.pkl', 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)

def load_obj(name):
	# load dictionaries so we can continue in middle of season
    with open('obj/' + name + '.pkl', 'rb') as f:
        return pickle.load(f)

def get_obp(table, batter_stats):
	# Calculate the average OBP of starting batters
	# and update OBP for all players
	# Inputs:
	#	table: bs4.element.Tag -- table of batters
	#	batter_stats: dictionary of Batter objects
	# Outputs:
	#	obp: float -- average OBP of stating players
	#	batter_stats: dictionary of Batter objects
	rows = table.findAll(lambda tag: tag.name == "tr")[1:-1]
	# (^ first row is labels of columns and last row is total ^)
	obp = 0.0
	for row in rows:
		# get batter name
		name = row.find(lambda tag: tag.name == "th").getText()
		if name == "":
			continue
		short_name = name.split() # remove position (e.g., "CF")
		short_name = "%s %s" % (short_name[0], short_name[1])
		# calculate average OBP of starting players
		if not (name[0].isspace()): # indicates starting player
			obp += batter_stats[short_name].obp
		# update batting average of all players who batted 
		indiv_obp = row.find("td", attrs = {"data-stat":"onbase_perc"}).getText()
		if indiv_obp != "":
			batter_stats[short_name].obp = float(indiv_obp)
	return obp / 9, batter_stats

def get_rbis(table):
	# get RBIs scored this game
	total_row = table.find("tfoot")
	rbis = int(total_row.find("td", attrs = {"data-stat":"RBI"}).getText())
	return rbis

def calculate_starting_era_whip(table, pitcher_stats):
	# calculate the starting pitcher's ERA and WHIP
	name_row = table.findAll(lambda tag: tag.name == "tr")[1]
	name = name_row.find(lambda tag: tag.name == "th").getText()
	short_name = name.split() # remove position (e.g. "CF")
	short_name = "%s %s" % (short_name[0], short_name[1])
	if short_name[-1] == ",":
		short_name = short_name[:-1]
	era = pitcher_stats[short_name].era()
	whip = pitcher_stats[short_name].whip()
	return era, whip

def update_pitchers(table, pitcher_stats):
	# update pitcher stats
	rows = table.findAll(lambda tag: tag.name == "tr")[1:-1]
	# (^ first row is labels of columns and last row is total ^)
	for row in rows:
		# get pitcher name
		name = row.find(lambda tag: tag.name == "th").getText()
		short_name = name.split() # remove position (e.g. "CF")
		short_name = "%s %s" % (short_name[0], short_name[1])
		if short_name[-1] == ",":
			short_name = short_name[:-1]
		# update individual statistics
		pitcher_stats[short_name].ip += float((row.find("td", attrs = {"data-stat":"IP"}).getText()))
		pitcher_stats[short_name].hits += float((row.find("td", attrs = {"data-stat":"H"}).getText()))
		pitcher_stats[short_name].er += float((row.find("td", attrs = {"data-stat":"ER"}).getText()))
		pitcher_stats[short_name].bb += float((row.find("td", attrs = {"data-stat":"BB"}).getText()))
	return pitcher_stats


def single_game(boxscore_url, pitcher_stats, batter_stats, team_stats):
	# Get training data for a single game
	# Inputs:
	# 	boxscore_url: url to append to base url to get to boxscore
	#	pitcher_stats: dictionary of Pitcher objects
	#	team_stats: dictionary of Team objects
	#
	# Outputs:
	#	input_vec: training data input
	#	ouptut_vec: training data output
	#	pitcher_stats: updated dictionary of Pitcher objects
	#	team_stats: updated dictionary of Team objects

	# load page HTML into parser
	full_url = base_url + boxscore_url
	response = get(full_url)
	# for some reason the tables are commented out, so remove html comment tags
	comm_tag = re.compile("<!--|-->")
	new_response_text = re.sub(comm_tag, "", response.text)
	game_soup = BeautifulSoup(new_response_text, "html5lib") # 'html.parser'? 'lxml'?

	# get score information
	scorebox = game_soup.find("div", class_ = "scorebox")
	# get the team names
	names = scorebox.findAll("strong")
	away_team = names[0].getText().strip()
	home_team = names[1].getText().strip()

	# number of team in alphabetical ordering of MLB teams
	away_id = get_id(away_team)
	home_id = get_id(home_team)

	print "Game: %s vs. %s" % (away_team, home_team)

	### 4. Average runs scored ###
	away_ars = team_stats[away_team].avg_runs()
	home_ars = team_stats[home_team].avg_runs()

	### 5. Team RBIs ###
	away_rbis = team_stats[away_team].rbis
	home_rbis = team_stats[home_team].rbis

	# get the score
	scores = scorebox.findAll("div", class_ = "score")
	away_team_runs = int(scores[0].getText())
	home_team_runs = int(scores[1].getText())

	# update team games and runs
	team_stats[away_team].games += 1
	team_stats[home_team].games += 1
	team_stats[away_team].runs += away_team_runs
	team_stats[home_team].runs += home_team_runs

	# create output vector
	if away_team_runs > home_team_runs:
		output_vec = [1, 0]
	elif away_team_runs < home_team_runs:
		output_vec = [0, 1]
	else:
		# just in case :)
		print "TIE: %s vs. %s" % (away_team, home_team)
		return None, None, pitcher_stats, batter_stats, team_stats

	# get batting/pitching information
	stat_tables = game_soup.findAll("div", class_ = "table_wrapper")
	away_batting = stat_tables[0]
	home_batting = stat_tables[1]
	away_pitching = stat_tables[2]
	home_pitching = stat_tables[3]

	### 3. Team OBP ###
	away_obp, batter_stats = get_obp(away_batting, batter_stats)
	home_obp, batter_stats = get_obp(home_batting, batter_stats)

	# update team RBIs
	team_stats[away_team].rbis += get_rbis(away_batting)
	team_stats[home_team].rbis += get_rbis(home_batting)

	### 1. Starting pitcher ERA ###
	# and
	### 2. Starting pitcher WHIP ###
	away_era, away_whip = calculate_starting_era_whip(away_pitching, pitcher_stats)
	home_era, home_whip = calculate_starting_era_whip(home_pitching, pitcher_stats)

	# update pitching stats
	pitcher_stats = update_pitchers(away_pitching, pitcher_stats)
	pitcher_stats = update_pitchers(home_pitching, pitcher_stats)

	# print team_stats[away_team]
	# print team_stats[home_team]
	# print pitcher_stats["Madison Bumgarner"]

	input_vec = [away_era, away_whip, away_obp, away_ars, away_rbis,
				 home_era, home_whip, home_obp, home_ars, home_rbis,
				 away_id, home_id]
	return [input_vec, output_vec, pitcher_stats, batter_stats, team_stats]

def single_game_players(boxscore_url, every_player = True):
	# return a list of all (if every_player == True)/starting players in a game
	# load page HTML into parser
	full_url = base_url + boxscore_url
	response = get(full_url)
	# for some reason the tables are commented out, so remove html comment tags
	comm_tag = re.compile("<!--|-->")
	new_response_text = re.sub(comm_tag, "", response.text)
	game_soup = BeautifulSoup(new_response_text, "html5lib") # 'html.parser'? 'lxml'?

	stat_tables = game_soup.findAll("div", class_ = "table_wrapper")
	# away players === stat_tables[0]
	# home players === stat_tables[1]

	away_players = []
	home_players = []
	players = [away_players, home_players]
	# this is janky but it makes duplicating pitcher easy
	for i in xrange(2):
		table = stat_tables[i]
		rows = table.findAll(lambda tag: tag.name == "tr")[1:-1]
		# (^ first row is labels of columns and last row is total ^)
		extra = "EXTRA"
		for row in rows:
			# get batter name
			name = row.find(lambda tag: tag.name == "th").getText()
			if name == "":
				continue
			broken_name = name.split() # remove position (e.g., "CF")
			short_name = "%s %s" % (broken_name[0], broken_name[1])
			if short_name[-1] == ",":
				short_name = short_name[:-1]
			if every_player == False: 
				if not (name[0].isspace()): # indicates starting player
					if broken_name[-1] == "P":
						extra = short_name
					players[i].append(short_name)
			else:
				players[i].append(short_name)
		if len(players[i]) == 9:
			if extra == "EXTRA":
				pitching_table = stat_tables[i + 2]
				p_rows = pitching_table.findAll(lambda tag: tag.name == "tr")[1:-1]
				starting_pitcher_name = p_rows[0].find(lambda tag: tag.name == "th").getText()
				broken_name = starting_pitcher_name.split() # remove position (e.g., "CF")
				short_name = "%s %s" % (broken_name[0], broken_name[1])
				if short_name[-1] == ",":
					short_name = short_name[:-1]
				players[i].append(short_name)
			else:
				players[i].append(extra)
			# here we are counting the pitcher twice for NL games
			# the graph needs to have a constant number of inputs
			# the pitcher seems like a good choice for a "tenth" player
	return away_players + home_players

def single_game_win_lose(boxscore_url):
	# get away win vector or home win vector for a single game
	# e.g., [1, 0] or [0, 1]

	# load page HTML into parser
	full_url = base_url + boxscore_url
	response = get(full_url)
	# for some reason the tables are commented out, so remove html comment tags
	comm_tag = re.compile("<!--|-->")
	new_response_text = re.sub(comm_tag, "", response.text)
	game_soup = BeautifulSoup(new_response_text, "html5lib") # 'html.parser'? 'lxml'?

	# get score information
	scorebox = game_soup.find("div", class_ = "scorebox")
	# get the score
	scores = scorebox.findAll("div", class_ = "score")
	away_team_runs = int(scores[0].getText())
	home_team_runs = int(scores[1].getText())

	# create output vector
	if away_team_runs > home_team_runs:
		return [1, 0]
	elif away_team_runs < home_team_runs:
		return [0, 1]
	else:
		# just in case :)
		print "TIE: %s vs. %s" % (away_team, home_team)
		return None, None

def main(mode):
	# div names
	# div_2015 = "div_9796507309"
	# div_2016 = "div_1992881718"
	# div_2017 = "div_8585337858"
	# div_2018 = "div_8577026772"

	# change "url" and "div_name" based on what year you want
	url = "https://www.baseball-reference.com/leagues/MLB/2015-schedule.shtml"
	div_name = "div_9796507309"
	response = get(url)
	html_soup = BeautifulSoup(response.text, "html.parser")
	html_soup = html_soup.find("div", {"id": div_name})
	gamedays = html_soup.findAll("div")

	if mode == 0:
	# collect data on teams
		# excel spreadsheet to write to
		wb = Workbook(write_only = True)
		dest_file = "training_data_players_15.xlsx"
		sheet = wb.create_sheet()

		pitcher_stats = keydefaultdict(Pitcher)
		batter_stats = keydefaultdict(Batter)
		team_stats = keydefaultdict(Team)

		game_count = 1
		for day in gamedays:
			games_today = day.find_all("p", class_ = "game")
			for game in games_today:
				print "Game %d" % game_count
				links = game.find_all("a", href = True)
				if links[2].getText() == "Preview":
					return
				boxscore = links[2]["href"] # 3rd link in game (away team, home team, boxscore)
				input_vec, output_vec, pitcher_stats, batter_stats, team_stats = single_game(boxscore, pitcher_stats, batter_stats, team_stats)
				if (input_vec is None) and (output_vec is None):
					print "skipped"
					continue
				else:
					training_example = input_vec + output_vec
					for num, value in enumerate(training_example):
						sheet.cell(column = num + 1, row = game_count, value = value)
					game_count += 1
			wb.save(filename = dest_file)
			save_obj(batter_stats, "batters")
			save_obj(pitcher_stats, "pitchers")
			save_obj(team_stats, "teams")
			# it's not really necessary to save these dicts, but why not

	elif mode == 1:
	# collect data on players
		# using csv instead of xlsx because of internet problems
		player_data_file = open("training_data_players_15.csv", "w")

		# all_players = [] # for first time runthrough collecting player names
		all_players = load_obj("all_players_list")
		le = preprocessing.LabelEncoder()
		le.fit(all_players)

		game_count = 0
		div_count = 0
		with player_data_file:
			writer = csv.writer(player_data_file)
			for day in gamedays[div_count:]:
				games_today = day.find_all("p", class_ = "game")
				for game in games_today:
					print "Game %d" % game_count
					links = game.find_all("a", href = True)
					if links[2].getText() == "Preview":
						save_obj(all_players, "all_players_list")
						return
			 		boxscore = links[2]["href"] # 3rd link in game (away team, home team, boxscore)
			 		game_output = single_game_win_lose(boxscore)
			 		if game_output is None:
			 			continue
			 		game_output = np.array(game_output)
			 		game_players = single_game_players(boxscore, every_player = False)
			 		# this stuff (below) is for first time run through of collecting player names
			 		# game_players = single_game_players(boxscore, every_player = True)
			 		# for player in game_players:
			 		# 	if player not in all_players:
			 		# 		all_players.append(player)
			 		transformed = np.array(le.transform(game_players))
			 		training_example = np.append(transformed, game_output)
			 		print "Training example:", training_example
					writer.writerow(training_example)
			 		game_count += 1
			 	div_count += 1
			 	print "Gameday:", div_count
			save_obj(all_players, "all_players_list")
	else:
		print "invalid mode"
		exit()

if __name__ == '__main__':
	if len(sys.argv) < 2:
		print "Usage: data_collection.py (-players | -teams)"
		exit()

	mode = -1
	if sys.argv[1] == "-teams":
		mode = 0
	elif sys.argv[1] == "-players":
		mode = 1
	else:
		print "Usage: data_collection.py (-players | -teams)"
		exit()

	main(mode)