from openpyxl import load_workbook
# from __future__ import print_function

import tensorflow as tf
import numpy as np
import pickle
import sys
	
# collect new data or load old data
collect = 0
if len(sys.argv) > 1 and sys.argv[1] == "-collect":
	collect = 1

# Parameters
learning_rate = 0.001
training_epochs = 80
batch_size = 10
display_step = 1

# Network Parameters
n_hidden_1 = 32 # hidden layer number of neurons
n_input = 60 # number of features in input 
n_classes = 2 # win or loss

# get data
# training data
if collect:
	num_examples = 7286 # number of training examples
	wb = load_workbook(filename = "training_data_teams_15_16_17.xlsx", read_only = True)
	ws_train = wb['Data']
	training_data = []
	for index, row in enumerate(ws_train.iter_rows(min_col = 11, max_col = 15)):
	# 15 above because of year value
		# print "training row %d" % index
		repeat = {2015:1, 2016:4, 2017:9}[row[-1].value]
		# print "repeating %d times" % repeat
		for _ in xrange(repeat):
			this_row = []
			for cell in row[:-1]:
				this_row.append(cell.value)
			training_data.append(this_row)
	training_data = np.asarray(training_data)
	with open('obj/'+ "training_data_15_16_17" + '.pkl', 'wb') as f:
		pickle.dump(training_data, f, pickle.HIGHEST_PROTOCOL)
else:
	with open('obj/' + "training_data_15_16_17" + '.pkl', 'rb') as f:
		training_data = pickle.load(f)

# testing data
if collect:
	num_tests = 654 # number of testing examples
	wb_test = load_workbook(filename = "training_data_teams_18.xlsx", read_only = True)
	ws_test = wb_test['Data']
	testing_data = []
	for index, row in enumerate(ws_test.iter_rows(min_col = 11, max_col = 14)):
		print "testing row %d" % index
		this_row = []
		for cell in row:
			print "test value", cell.value
			this_row.append(cell.value)
		testing_data.append(this_row)
	testing_data = np.asarray(testing_data)
	with open('obj/'+ "test_data_teams_18" + '.pkl', 'wb') as f:
		pickle.dump(testing_data, f, pickle.HIGHEST_PROTOCOL)
else:
	with open('obj/' + "test_data_teams_18" + '.pkl', 'rb') as f:
		testing_data = pickle.load(f)

###
# training_data = training_data[4856:] # only 2017 data
###

# tf stuff
initializer = tf.contrib.layers.xavier_initializer()
embedded_teams = tf.Variable(initializer([30, 30]))	
# tf Graph input
t1 = tf.placeholder(tf.int32) #, shape=[batch_size, n_input/2]) # away team embedding
t2 = tf.placeholder(tf.int32) #, shape=[batch_size, n_input/2]) # home team embedding
X1 = tf.nn.embedding_lookup(embedded_teams, t1)
X2 = tf.nn.embedding_lookup(embedded_teams, t2)
print "X1:", X1
print "X2:", X2
X = tf.concat([X1, X2], axis = 1)
print "X:", X
Y = tf.placeholder(tf.int32, [None, n_classes])
keep_prob = tf.placeholder(tf.float32)

# Store layers weight & bias
weights = {
    'h1': tf.Variable(initializer([n_input, n_hidden_1])),
    'out': tf.Variable(initializer([n_hidden_1, n_classes]))
}
biases = {
    'b1': tf.Variable(initializer([n_hidden_1])),
    'out': tf.Variable(initializer([n_classes]))
}

# Create model
def multilayer_perceptron(x, keep_prob):
    # Hidden fully connected layer with 256 neurons
    layer_1 = tf.add(tf.matmul(x, weights['h1']), biases['b1'])
    # Output fully connected layer with a neuron for each class
    dropout = tf.nn.dropout(layer_1, keep_prob)
    out_layer = tf.matmul(layer_1, weights['out']) + biases['out']
    return out_layer

# Construct model
logits = multilayer_perceptron(X, keep_prob)

# Define loss and optimizer
loss_op = tf.reduce_mean(tf.nn.softmax_cross_entropy_with_logits(
    logits = logits, labels = Y))
optimizer = tf.train.AdamOptimizer(learning_rate = learning_rate)
train_op = optimizer.minimize(loss_op)
# Initializing the variables
init = tf.global_variables_initializer()

with tf.Session() as sess:
	sess.run(init)

	np.random.shuffle(training_data)

    # Training cycle
	for epoch in xrange(training_epochs):
		avg_cost = 0.
		total_batch = int(training_data.shape[0]/batch_size)
		print "total_batch", total_batch
		shuffled = training_data
		np.random.shuffle(shuffled)
		# Loop over all batches
		for i in xrange(total_batch):
		    batch_t1 = []
		    batch_t2 = []
		    batch_y = []
		    start_index = i * batch_size
		    for j in xrange(start_index, start_index + batch_size):
		    	datum = shuffled[j]
		    	# print "datum %d" % j, datum
		    	# print "appending", tf.nn.embedding_lookup(teams, [datum[0]]), "to batch_t1"
		    	batch_t1.append(int(datum[0]))
		    	batch_t2.append(int(datum[1]))
		    	batch_y.append([int(datum[2]), int(datum[3])])
		    
		    # Run optimization op (backprop) and cost op (to get loss value)
		    _, c = sess.run([train_op, loss_op], feed_dict = {t1: batch_t1,
		    												  t2: batch_t2,
		                                                      Y: batch_y,
		                                                      keep_prob: .9})
		    # Compute average loss
		    avg_cost += c / total_batch
		# Display logs per epoch step
		if epoch % display_step == 0:
		    print("Epoch:", '%04d' % (epoch+1), "cost={:.9f}".format(avg_cost))
	print("Optimization Finished!")

	# format test data
	test_t1 = []
	test_t2 = []
	test_y = []
	for datum in testing_data:
		test_t1.append(datum[0])
		test_t2.append(datum[1])
		test_y.append([datum[2], datum[3]])

	# Test model
	pred = tf.nn.softmax(logits)  # Apply softmax to logits
	correct_prediction = tf.equal(tf.argmax(pred, 1), tf.argmax(Y, 1))
	# Calculate accuracy
	accuracy = tf.reduce_mean(tf.cast(correct_prediction, "float"))
	print("Accuracy:", accuracy.eval({t1: test_t1, t2: test_t2, Y: test_y, keep_prob: 1.0}))